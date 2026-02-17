# -*- coding: utf-8 -*-
"""
Деобезличивание данных в data3.xlsx.

Алгоритм: шифр Цезаря с индивидуальным ключом для каждой строки.
Ключ находится по адресу (перебор сдвигов, выбор по правдоподобности).
Для латиницы (email) используется эквивалентный сдвиг: (33 - ключ) % 26.
"""
import openpyxl
from openpyxl import Workbook

RUS_ALPHABET = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'
ADDR_MARKERS = ['ул.', 'д.', 'кв.', 'пр.', 'г.', 'обл.', 'стр.', 'корп.', 'пл.', 'пер.']


def decrypt_cyrillic(text, key):
    if not text or not isinstance(text, str):
        return text
    result = []
    for c in text:
        if c in RUS_ALPHABET:
            pos = RUS_ALPHABET.index(c)
            result.append(RUS_ALPHABET[(pos + key) % 33])
        elif c in RUS_ALPHABET.upper():
            pos = RUS_ALPHABET.upper().index(c)
            result.append(RUS_ALPHABET[(pos + key) % 33].upper())
        else:
            result.append(c)
    return ''.join(result)


def decrypt_latin(text, key):
    if not text or not isinstance(text, str):
        return text
    shift = key % 26
    result = []
    for c in text:
        if 'a' <= c <= 'z':
            result.append(chr((ord(c) - ord('a') - shift) % 26 + ord('a')))
        elif 'A' <= c <= 'Z':
            result.append(chr((ord(c) - ord('A') - shift) % 26 + ord('A')))
        elif '0' <= c <= '9':
            result.append(c)
        else:
            result.append(c)
    return ''.join(result)


def score_address(text):
    if not text:
        return 0
    score = 0
    for m in ADDR_MARKERS:
        if m in text:
            score += 2
    if text.strip().startswith('ул.'):
        score += 3
    for c in 'аеинорст':
        score += text.lower().count(c) * 0.1
    return score


def find_key(address):
    if not address:
        return 0
    best_key = 0
    best_score = -1
    for key in range(33):
        dec = decrypt_cyrillic(address, key)
        sc = score_address(dec)
        if sc > best_score:
            best_score = sc
            best_key = key
    return best_key


def main():
    wb = openpyxl.load_workbook('data3.xlsx', read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    data_rows = [r for r in all_rows[2:] if (r[1] or r[2] or r[3])]

    result = Workbook()
    rs = result.active
    rs.append(['Телефон', 'email', 'Адрес', 'Ключ'])

    for row in data_rows:
        phone_hash = row[1] or ''
        email_enc = row[2] or ''
        addr_enc = row[3] or ''

        key = find_key(addr_enc)
        addr_dec = decrypt_cyrillic(addr_enc, key)
        email_shift = (33 - key) % 26
        email_dec = decrypt_latin(email_enc, email_shift)

        rs.append([phone_hash, email_dec, addr_dec, key])

    result.save('data3_deobfuscated.xlsx')
    print(f'Обработано строк: {len(data_rows)}')
    print('Результат сохранён в data3_deobfuscated.xlsx')


if __name__ == '__main__':
    main()
